#!/usr/bin/env python3
"""Preview rows from a specific worksheet for xls/xlsx files."""

from __future__ import annotations

import argparse
from pathlib import Path


def normalize_cell(value: object) -> str:
    if value is None:
        return ""
    return str(value)


def preview_xls(path: Path, sheet_name: str | None, sheet_index: int | None, max_rows: int) -> tuple[str, int, int, list[list[str]]]:
    import xlrd

    workbook = xlrd.open_workbook(str(path))
    if sheet_name is not None:
        sheet = workbook.sheet_by_name(sheet_name)
    elif sheet_index is not None:
        sheet = workbook.sheet_by_index(sheet_index)
    else:
        sheet = workbook.sheet_by_index(0)

    rows = []
    for row_index in range(min(sheet.nrows, max_rows)):
        rows.append([normalize_cell(v) for v in sheet.row_values(row_index)])
    return sheet.name, sheet.nrows, sheet.ncols, rows


def preview_xlsx(path: Path, sheet_name: str | None, sheet_index: int | None, max_rows: int) -> tuple[str, int, int, list[list[str]]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name is not None:
        sheet = workbook[sheet_name]
    elif sheet_index is not None:
        sheet = workbook[workbook.sheetnames[sheet_index]]
    else:
        sheet = workbook[workbook.sheetnames[0]]

    rows = []
    for index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if index > max_rows:
            break
        rows.append([normalize_cell(v) for v in row])
    return sheet.title, sheet.max_row, sheet.max_column, rows


def main() -> int:
    parser = argparse.ArgumentParser(description="Preview one worksheet from an xls/xlsx file.")
    parser.add_argument("path", help="Path to xls/xlsx file")
    group = parser.add_mutually_exclusive_group()
    group.add_argument("--sheet", help="Worksheet name")
    group.add_argument("--index", type=int, help="Worksheet index (0-based)")
    parser.add_argument("--rows", type=int, default=40, help="Max rows to print")
    args = parser.parse_args()

    path = Path(args.path).expanduser()
    suffix = path.suffix.lower()

    if suffix == ".xls":
        title, total_rows, total_cols, rows = preview_xls(path, args.sheet, args.index, args.rows)
    elif suffix == ".xlsx":
        title, total_rows, total_cols, rows = preview_xlsx(path, args.sheet, args.index, args.rows)
    else:
        raise SystemExit(f"Unsupported workbook format: {path.suffix}")

    print(f"[sheet] {title}")
    print(f"[size] rows={total_rows} cols={total_cols}")
    print(f"[preview_rows] {min(total_rows, args.rows)}")
    for row in rows:
        print("\t".join(row))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
