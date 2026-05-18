#!/usr/bin/env python3
"""Inspect workbook structure for xls/xlsx files."""

from __future__ import annotations

import argparse
import json
from pathlib import Path


def inspect_xls(path: Path) -> dict:
    import xlrd

    workbook = xlrd.open_workbook(str(path))
    sheets = []
    for index in range(workbook.nsheets):
        sheet = workbook.sheet_by_index(index)
        sheets.append(
            {
                "index": index,
                "name": sheet.name,
                "rows": sheet.nrows,
                "cols": sheet.ncols,
            }
        )
    return {
        "path": str(path),
        "format": "xls",
        "sheet_count": workbook.nsheets,
        "sheets": sheets,
    }


def inspect_xlsx(path: Path) -> dict:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheets = []
    for index, name in enumerate(workbook.sheetnames):
        sheet = workbook[name]
        sheets.append(
            {
                "index": index,
                "name": name,
                "rows": sheet.max_row,
                "cols": sheet.max_column,
            }
        )
    return {
        "path": str(path),
        "format": "xlsx",
        "sheet_count": len(workbook.sheetnames),
        "sheets": sheets,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Inspect workbook sheets for xls/xlsx files.")
    parser.add_argument("path", help="Path to xls/xlsx file")
    parser.add_argument("--json", action="store_true", help="Output JSON instead of plain text")
    args = parser.parse_args()

    path = Path(args.path).expanduser()
    suffix = path.suffix.lower()

    if suffix == ".xls":
        result = inspect_xls(path)
    elif suffix == ".xlsx":
        result = inspect_xlsx(path)
    else:
        raise SystemExit(f"Unsupported workbook format: {path.suffix}")

    if args.json:
        print(json.dumps(result, ensure_ascii=False, indent=2))
        return 0

    print(f"文件: {result['path']}")
    print(f"格式: {result['format']}")
    print(f"工作表数量: {result['sheet_count']}")
    for sheet in result["sheets"]:
        print(f"{sheet['index']}\t{sheet['name']}\t{sheet['rows']}\t{sheet['cols']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
