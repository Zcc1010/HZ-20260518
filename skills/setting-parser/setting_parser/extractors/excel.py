"""Excel 提取器：openpyxl 读取所有 sheet."""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

from .base import ExtractedText


class ExcelExtractedText(ExtractedText):
    """ExtractedText 的 Excel 专用子类，保留按 sheet 名索引的表格视图."""

    def __init__(
        self,
        source_path: str,
        tables_by_sheet: dict[str, list[list[str]]],
        warnings: Optional[list[str]] = None,
    ):
        # 初始化父类字段（绕开 dataclass __init__ 的字段顺序限制）
        self.source_path = source_path
        self.markdown = self._sheets_to_markdown(tables_by_sheet)
        self.tables = list(tables_by_sheet.values())
        self.warnings = warnings or []
        # 子类扩展字段
        self.tables_by_sheet = tables_by_sheet

    @staticmethod
    def _sheets_to_markdown(tables_by_sheet: dict[str, list[list[str]]]) -> str:
        parts: list[str] = []
        for sheet_name, rows in tables_by_sheet.items():
            parts.append(f"## {sheet_name}\n")
            for row in rows:
                parts.append("| " + " | ".join(str(c) for c in row) + " |")
            parts.append("")
        return "\n".join(parts)


def extract_excel(path: str) -> ExcelExtractedText:
    """从 Excel 提取所有 sheet 的表格."""
    warnings: list[str] = []
    if not Path(path).exists():
        return ExcelExtractedText(
            source_path=path,
            tables_by_sheet={},
            warnings=[f"文件不存在: {path}"],
        )

    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True, read_only=True)
    tables: dict[str, list[list[str]]] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows: list[list[str]] = []
        for row in ws.iter_rows(values_only=True):
            rows.append(["" if c is None else str(c) for c in row])
        if rows:
            tables[sheet_name] = rows

    return ExcelExtractedText(source_path=path, tables_by_sheet=tables, warnings=warnings)