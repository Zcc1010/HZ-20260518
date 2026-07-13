# -*- coding: utf-8 -*-
"""定值校核 V2 API — 工作区/文件管理（纯 Python 实现，不依赖外部服务）"""
from __future__ import annotations

import asyncio
import base64
import io
import json
import logging
import os
import re
import shutil
import subprocess
import time
from pathlib import Path
from typing import Any

from fastapi import APIRouter, Request, UploadFile
from fastapi.responses import FileResponse, Response, StreamingResponse

logger = logging.getLogger(__name__)

router = APIRouter()

# ── 工作区根目录 ──

def _workspace_root() -> Path:
    """~/.nanobot/agentplayground/setting-check/workspace/"""
    home = Path.home()
    root = home / ".nanobot" / "agentplayground" / "setting-check" / "workspace"
    root.mkdir(parents=True, exist_ok=True)
    return root


def _ws_path(ws: str) -> Path:
    """获取工作区绝对路径"""
    return _workspace_root() / ws


def _safe_join(base: Path, target: str) -> Path:
    """防止路径穿越，允许软链接指向 resources 和 temp 目录"""
    result = (base / target).resolve()
    base_resolved = base.resolve()
    # 允许访问工作区内的文件
    if str(result).startswith(str(base_resolved)):
        return result
    # 允许访问 agentplayground/resources 目录下的文件（说明书等）
    resources_dir = (Path.home() / ".nanobot" / "agentplayground" / "resources").resolve()
    if str(result).startswith(str(resources_dir)):
        return result
    # 允许访问 agentplayground/temp 目录下的文件（临时筛选的整定原则和台账）
    temp_dir = (Path.home() / ".nanobot" / "agentplayground" / "temp").resolve()
    if str(result).startswith(str(temp_dir)):
        return result
    raise ValueError("path traversal detected")


# ── 文件树构建 ──

IMAGE_EXTS = {"png", "jpg", "jpeg", "gif", "svg", "webp"}
BINARY_EXTS = {"xls", "xlsx", "docx"}
PDF_EXTS = {"pdf"}
EXCLUDED_DIRS = {"node_modules", "temp", "__pycache__", ".git"}


def _build_tree(dir_path: Path, base_path: Path) -> list[dict[str, Any]]:
    """递归构建文件树"""
    items = []
    if not dir_path.exists() or not dir_path.is_dir():
        return items

    for entry in sorted(dir_path.iterdir(), key=lambda e: (not e.is_dir(), e.name.lower())):
        if entry.name.startswith(".") or entry.name in EXCLUDED_DIRS:
            continue
        stat = entry.stat()
        rel_path = str(entry.relative_to(base_path)).replace("\\", "/")
        node: dict[str, Any] = {
            "name": entry.name,
            "path": rel_path,
            "type": "directory" if entry.is_dir() else "file",
            "size": stat.st_size,
            "mtime": stat.st_mtime,
        }
        if entry.is_dir():
            node["children"] = _build_tree(entry, base_path)
            # 报告目录只保留 .md 文件
            if entry.name == "报告":
                node["children"] = [c for c in node["children"] if c["type"] == "directory" or c["name"].endswith(".md")]
        items.append(node)

    # 报告目录排到最后
    items.sort(key=lambda e: (e["name"] == "报告", not e["type"] == "directory", e["name"].lower()))
    return items


# ── 工作区 CRUD ──

@router.get("/workspaces")
async def list_workspaces():
    root = _workspace_root()
    items = []
    for entry in sorted(root.iterdir(), key=lambda e: e.name):
        if entry.is_dir() and not entry.name.startswith(".") and entry.name not in EXCLUDED_DIRS:
            items.append({"name": entry.name, "type": "directory"})
    return {"items": items}


@router.post("/workspaces")
async def create_workspace(request: Request):
    body = await request.json()
    name = body.get("name", "").strip()
    if not name or "/" in name or "\\" in name:
        return Response(content=json.dumps({"error": "invalid name"}), status_code=400, media_type="application/json")
    target = _ws_path(name)
    if target.exists():
        return Response(content=json.dumps({"error": "already exists"}), status_code=409, media_type="application/json")
    target.mkdir(parents=True, exist_ok=True)
    return {"name": name}


@router.patch("/workspaces/{ws}")
async def rename_workspace(ws: str, request: Request):
    body = await request.json()
    new_name = body.get("name", "").strip()
    if not new_name or "/" in new_name or "\\" in new_name:
        return Response(content=json.dumps({"error": "invalid name"}), status_code=400, media_type="application/json")
    old_path = _ws_path(ws)
    new_path = _ws_path(new_name)
    if not old_path.exists():
        return Response(content=json.dumps({"error": "not found"}), status_code=404, media_type="application/json")
    if new_path.exists():
        return Response(content=json.dumps({"error": "name exists"}), status_code=409, media_type="application/json")
    old_path.rename(new_path)
    return {"name": new_name}


@router.delete("/workspaces/{ws}")
async def delete_workspace(ws: str):
    target = _ws_path(ws)
    if not target.exists():
        return Response(content=json.dumps({"error": "not found"}), status_code=404, media_type="application/json")
    shutil.rmtree(target)
    return {"ok": True}


# ── 文件操作 ──

@router.get("/workspaces/{ws}/tree")
async def get_file_tree(ws: str):
    dir_path = _ws_path(ws)
    if not dir_path.exists() or not dir_path.is_dir():
        return Response(content=json.dumps({"error": "not found"}), status_code=404, media_type="application/json")
    return _build_tree(dir_path, dir_path)


@router.get("/workspaces/{ws}/search")
async def search_files(ws: str, q: str = ""):
    dir_path = _ws_path(ws)
    if not dir_path.exists():
        return {"items": []}
    q_lower = q.lower()
    results = []
    for root, _, files in os.walk(dir_path):
        for f in files:
            if f.startswith("."):
                continue
            if q_lower in f.lower():
                full = Path(root) / f
                results.append({"name": f, "path": str(full.relative_to(dir_path)).replace("\\", "/")})
    return {"items": results}


@router.get("/workspaces/{ws}/read")
async def read_file(ws: str, path: str = ""):
    if not path:
        return Response(content=json.dumps({"error": "missing path"}), status_code=400, media_type="application/json")
    try:
        full = _safe_join(_ws_path(ws), path)
    except ValueError:
        return Response(content=json.dumps({"error": "path traversal"}), status_code=400, media_type="application/json")
    if not full.exists():
        return Response(content=json.dumps({"error": "not found"}), status_code=404, media_type="application/json")
    if full.is_dir():
        return Response(content=json.dumps({"error": "is directory"}), status_code=400, media_type="application/json")

    ext = full.suffix.lower().lstrip(".")

    # 图片：直接返回二进制
    if ext in IMAGE_EXTS:
        mime = f"image/{'svg+xml' if ext == 'svg' else ext}"
        return FileResponse(str(full), media_type=mime)

    # PDF：返回二进制
    if ext in PDF_EXTS:
        return FileResponse(str(full), media_type="application/pdf", headers={"Content-Disposition": "inline"})

    # doc：转 HTML
    if ext == "doc":
        try:
            result = subprocess.run(
                ["antiword", str(full)],
                capture_output=True, text=True, timeout=30,
            )
            if result.returncode == 0:
                text = result.stdout
                html = "".join(f"<p>{line}</p>" for line in text.split("\n") if line.strip())
                return Response(content=html, media_type="text/html")
        except Exception:
            pass
        return Response(content=json.dumps({"error": "doc 转换失败"}), status_code=500, media_type="application/json")

    # 二进制格式（xlsx/docx）：返回 base64
    if ext in BINARY_EXTS:
        data = full.read_bytes()
        return {"base64": base64.b64encode(data).decode(), "name": full.name, "ext": ext}

    # 文本：直接返回
    try:
        text = full.read_text(encoding="utf-8")
    except UnicodeDecodeError:
        text = full.read_text(encoding="gbk", errors="replace")
    return Response(content=text, media_type="text/plain; charset=utf-8")


@router.put("/workspaces/{ws}/write")
async def write_file(ws: str, request: Request):
    body = await request.json()
    path = body.get("path", "")
    content = body.get("content", "")
    if not path:
        return Response(content=json.dumps({"error": "missing path"}), status_code=400, media_type="application/json")
    try:
        full = _safe_join(_ws_path(ws), path)
    except ValueError:
        return Response(content=json.dumps({"error": "path traversal"}), status_code=400, media_type="application/json")
    full.parent.mkdir(parents=True, exist_ok=True)
    full.write_text(content, encoding="utf-8")
    return {"ok": True}


@router.post("/workspaces/{ws}/rename")
async def rename_file(ws: str, request: Request):
    body = await request.json()
    path = body.get("path", "")
    new_name = body.get("newName", "")
    if not path or not new_name or "/" in new_name:
        return Response(content=json.dumps({"error": "invalid params"}), status_code=400, media_type="application/json")
    try:
        full = _safe_join(_ws_path(ws), path)
        dest = full.parent / new_name
    except ValueError:
        return Response(content=json.dumps({"error": "path traversal"}), status_code=400, media_type="application/json")
    if not full.exists():
        return Response(content=json.dumps({"error": "not found"}), status_code=404, media_type="application/json")
    full.rename(dest)
    return {"ok": True}


@router.post("/workspaces/{ws}/copy")
async def copy_file(ws: str, request: Request):
    body = await request.json()
    src = body.get("src", "")
    dest = body.get("dest", "")
    if not src or not dest:
        return Response(content=json.dumps({"error": "missing src or dest"}), status_code=400, media_type="application/json")
    try:
        src_full = _safe_join(_ws_path(ws), src)
        dest_full = _safe_join(_ws_path(ws), dest)
    except ValueError:
        return Response(content=json.dumps({"error": "path traversal"}), status_code=400, media_type="application/json")
    if not src_full.exists():
        return Response(content=json.dumps({"error": "src not found"}), status_code=404, media_type="application/json")
    dest_full.parent.mkdir(parents=True, exist_ok=True)
    if src_full.is_dir():
        shutil.copytree(src_full, dest_full)
    else:
        shutil.copy2(src_full, dest_full)
    return {"ok": True}


@router.post("/workspaces/{ws}/duplicate")
async def duplicate_file(ws: str, request: Request):
    body = await request.json()
    path = body.get("path", "")
    if not path:
        return Response(content=json.dumps({"error": "missing path"}), status_code=400, media_type="application/json")
    try:
        src_full = _safe_join(_ws_path(ws), path)
    except ValueError:
        return Response(content=json.dumps({"error": "path traversal"}), status_code=400, media_type="application/json")
    if not src_full.exists():
        return Response(content=json.dumps({"error": "not found"}), status_code=404, media_type="application/json")

    parent = src_full.parent
    ext = src_full.suffix
    base = src_full.stem

    dest_full = None
    for i in range(1, 100):
        candidate = parent / f"{base} 副本{i}{ext}"
        if not candidate.exists():
            dest_full = candidate
            break
    if not dest_full:
        return Response(content=json.dumps({"error": "too many copies"}), status_code=400, media_type="application/json")

    if src_full.is_dir():
        shutil.copytree(src_full, dest_full)
    else:
        shutil.copy2(src_full, dest_full)
    return {"ok": True}


@router.post("/workspaces/{ws}/move")
async def move_file(ws: str, request: Request):
    body = await request.json()
    src = body.get("src", "")
    dest = body.get("dest", "")
    if not src or not dest:
        return Response(content=json.dumps({"error": "missing src or dest"}), status_code=400, media_type="application/json")
    try:
        src_full = _safe_join(_ws_path(ws), src)
        dest_full = _safe_join(_ws_path(ws), dest)
    except ValueError:
        return Response(content=json.dumps({"error": "path traversal"}), status_code=400, media_type="application/json")
    if not src_full.exists():
        return Response(content=json.dumps({"error": "src not found"}), status_code=404, media_type="application/json")
    dest_full.parent.mkdir(parents=True, exist_ok=True)
    shutil.move(str(src_full), str(dest_full))
    return {"ok": True}


@router.delete("/workspaces/{ws}/file")
async def delete_file(ws: str, path: str = ""):
    if not path:
        return Response(content=json.dumps({"error": "missing path"}), status_code=400, media_type="application/json")
    try:
        full = _safe_join(_ws_path(ws), path)
    except ValueError:
        return Response(content=json.dumps({"error": "path traversal"}), status_code=400, media_type="application/json")
    if not full.exists():
        return Response(content=json.dumps({"error": "not found"}), status_code=404, media_type="application/json")
    if full.is_dir():
        shutil.rmtree(full)
    else:
        # 删除报告 .md 时同步删除对应的 .docx，反之亦然
        if full.suffix == ".md":
            docx = full.with_suffix(".docx")
            if docx.exists():
                docx.unlink()
        elif full.suffix == ".docx":
            md = full.with_suffix(".md")
            if md.exists():
                md.unlink()
        full.unlink()
    return {"ok": True}


@router.post("/workspaces/{ws}/upload")
async def upload_files(ws: str, request: Request):
    dir_path = _ws_path(ws)
    if not dir_path.exists():
        return Response(content=json.dumps({"error": "workspace not found"}), status_code=404, media_type="application/json")

    form = await request.form()
    uploaded = []

    for key, value in form.multi_items():
        if hasattr(value, "read"):  # UploadFile
            file_name = value.filename or "unnamed"
            content = await value.read()
            # 支持 category/filename 格式（前端上传时带分类前缀）
            try:
                dest = _safe_join(dir_path, file_name)
            except ValueError:
                continue
            dest.parent.mkdir(parents=True, exist_ok=True)
            dest.write_bytes(content)
            uploaded.append(file_name)

    return {"files": uploaded}


# ── SSE 文件变更事件 ──

def _get_dir_mtime(dir_path: Path) -> float:
    """获取目录下所有文件的最新修改时间"""
    max_mtime = 0.0
    if not dir_path.exists():
        return max_mtime
    for root, _, files in os.walk(dir_path):
        for f in files:
            try:
                mtime = os.path.getmtime(os.path.join(root, f))
                if mtime > max_mtime:
                    max_mtime = mtime
            except OSError:
                pass
    return max_mtime


@router.post("/workspaces/{ws}/link-manuals")
async def link_manuals(ws: str):
    """扫描工作区定值单，提取设备型号，查找并软链接说明书"""
    from webui.services.setting_check.device_extractor import DEVICE_EXTRACTOR_PROMPT

    dir_path = _ws_path(ws)
    if not dir_path.exists():
        return Response(content=json.dumps({"error": "workspace not found"}), status_code=404, media_type="application/json")

    # 1. 扫描定值单文件
    setting_dir = dir_path / "定值单"
    if not setting_dir.exists():
        return {"linked": False, "reason": "定值单目录不存在"}

    setting_files = []
    for f in sorted(setting_dir.iterdir()):
        if f.is_file() and f.suffix.lower() in {'.xls', '.xlsx', '.doc', '.docx', '.pdf', '.md', '.txt'}:
            setting_files.append(f)

    if not setting_files:
        return {"linked": False, "reason": "定值单目录为空"}

    # 2. 从文件名和内容提取设备信息
    # 简单提取：从工作区名和文件名中推测厂家和型号
    workspace_name = ws

    # manuals 目录
    manuals_root = Path.home() / ".nanobot" / "agentplayground" / "resources" / "manuals"
    logger.info(f"link_manuals: manuals_root={manuals_root}, exists={manuals_root.exists()}")

    # 收集所有可用的厂家/设备类型/型号
    available = {}  # {(manufacturer, device_type, model): path}
    if manuals_root.exists():
        for manufacturer_dir in manuals_root.iterdir():
            if not manufacturer_dir.is_dir() or manufacturer_dir.name.startswith(('.', '_')):
                continue
            manufacturer = manufacturer_dir.name
            for device_type_dir in manufacturer_dir.iterdir():
                if not device_type_dir.is_dir() or device_type_dir.name.startswith(('.', '_')):
                    continue
                device_type = device_type_dir.name
                for model_dir in device_type_dir.iterdir():
                    if not model_dir.is_dir() or model_dir.name.startswith(('.', '_')):
                        continue
                    model = model_dir.name
                    available[(manufacturer, device_type, model)] = model_dir

    if not available:
        logger.warning(f"link_manuals: no manuals available")
        return {"linked": False, "reason": "说明书资源为空"}

    # 3. 尝试从工作区名和文件名匹配
    # 读取前几个定值单文件的内容来提取信息
    content_samples = []
    for sf in setting_files[:3]:
        try:
            suffix = sf.suffix.lower()
            if suffix in {'.md', '.txt'}:
                text = sf.read_text(encoding="utf-8", errors="ignore")[:2000]
                content_samples.append(text)
            elif suffix in {'.docx', '.doc'}:
                try:
                    from docx import Document
                    doc = Document(str(sf))
                    text = "\n".join(p.text for p in doc.paragraphs)[:2000]
                    content_samples.append(text)
                except Exception:
                    pass
            elif suffix in {'.xlsx', '.xls'}:
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(str(sf), read_only=True, data_only=True)
                    text = ""
                    for sheet in wb.sheetnames:
                        ws = wb[sheet]
                        for row in ws.iter_rows(max_row=30, values_only=True):
                            text += " ".join(str(c) for c in row if c) + "\n"
                    content_samples.append(text[:2000])
                except Exception:
                    pass
        except Exception:
            pass

    all_text = workspace_name + " " + " ".join(f.name for f in setting_files)
    if content_samples:
        all_text += " " + " ".join(content_samples)

    # 标准化文本（全角转半角，去空格，统一大写）
    def normalize(s):
        s = s.replace("－", "-").replace("—", "-").replace("（", "(").replace("）", ")")
        s = s.replace("　", " ").strip()
        return s

    all_text_norm = normalize(all_text).upper()
    logger.info(f"link_manuals: available manufacturers={list(set(k[0] for k in available.keys()))}")
    logger.info(f"link_manuals: all_text[:200]={all_text[:200]}")

    # 尝试匹配厂家（精确匹配 + 前缀匹配 + 别名匹配）
    # 常见厂家别名映射
    MANUFACTURER_ALIASES = {
        "南瑞继保": ["南瑞", "NR", "NARI", "NRPC"],
        "南瑞科技": ["南瑞", "NR", "NARI"],
        "许继": ["XJ", "许继电气"],
        "四方继保": ["四方", "SF", "CSC"],
        "国电南自": ["南自", "SAC"],
        "长园深瑞": ["深瑞", "CY", "CYSR"],
        "北京四方": ["四方", "SF"],
    }

    matched_manufacturer = None
    manufacturers = list(set(k[0] for k in available))

    # 第一轮：精确子串匹配
    for mfr in manufacturers:
        if mfr in all_text:
            matched_manufacturer = mfr
            logger.info(f"link_manuals: exact matched manufacturer={mfr}")
            break

    # 第二轮：厂家名包含匹配（如"南瑞"匹配"南瑞继保"）
    if not matched_manufacturer:
        for mfr in manufacturers:
            mfr_norm = normalize(mfr).upper()
            if mfr_norm in all_text_norm or all_text_norm in mfr_norm:
                matched_manufacturer = mfr
                logger.info(f"link_manuals: prefix matched manufacturer={mfr}")
                break

    # 第三轮：别名匹配
    if not matched_manufacturer:
        for mfr, aliases in MANUFACTURER_ALIASES.items():
            if mfr in manufacturers:
                for alias in aliases:
                    if alias.upper() in all_text_norm:
                        matched_manufacturer = mfr
                        logger.info(f"link_manuals: alias matched manufacturer={mfr} via alias={alias}")
                        break
            if matched_manufacturer:
                break

    # 尝试匹配型号（精确匹配 + 前缀匹配）
    matched_model = None
    matched_device_type = None
    if matched_manufacturer:
        candidates = [(dt, mdl) for (mfr, dt, mdl) in available if mfr == matched_manufacturer]
        logger.info(f"link_manuals: candidates for {matched_manufacturer}={[(dt, mdl) for dt, mdl in candidates]}")

        # 第一轮：精确子串匹配
        for dt, mdl in candidates:
            if mdl in all_text:
                matched_model = mdl
                matched_device_type = dt
                logger.info(f"link_manuals: exact matched model={mdl}, device_type={dt}")
                break

        # 第二轮：型号前缀匹配（如"RCS-978"匹配"RCS-978A"）
        if not matched_model:
            for dt, mdl in candidates:
                mdl_norm = normalize(mdl).upper()
                text_norm = all_text_norm.replace("-", "").replace(" ", "")
                mdl_norm_clean = mdl_norm.replace("-", "").replace(" ", "")
                # 型号在文本中是前缀（如 RCS-978 在 RCS-978A 中）
                if len(mdl_norm_clean) >= 4:
                    for i in range(len(text_norm) - len(mdl_norm_clean) + 1):
                        if text_norm[i:i+len(mdl_norm_clean)] == mdl_norm_clean:
                            # 检查后面是否是字母或数字（型号后缀）
                            end_pos = i + len(mdl_norm_clean)
                            if end_pos >= len(text_norm) or not text_norm[end_pos].isalnum():
                                matched_model = mdl
                                matched_device_type = dt
                                logger.info(f"link_manuals: prefix matched model={mdl}, device_type={dt}")
                                break
                if matched_model:
                    break

    # 如果没匹配到型号，尝试跨厂家宽松匹配
    if not matched_model:
        logger.info(f"link_manuals: trying fuzzy match")
        for (mfr, dt, mdl) in available:
            model_clean = mdl.replace("-", "").replace(" ", "").upper()
            text_clean = all_text_norm.replace("-", "").replace(" ", "")
            if len(model_clean) >= 4 and model_clean in text_clean:
                matched_manufacturer = mfr
                matched_model = mdl
                matched_device_type = dt
                logger.info(f"link_manuals: fuzzy matched manufacturer={mfr}, model={mdl}, device_type={dt}")
                break

    if not matched_manufacturer or not matched_model:
        logger.warning(f"link_manuals: matching failed - manufacturer={matched_manufacturer}, model={matched_model}")

    # 4. 创建说明书软链接（如果匹配到）
    manual_linked = False
    manual_link_name = ""
    if matched_manufacturer and matched_model:
        manual_src = available.get((matched_manufacturer, matched_device_type, matched_model))
        if manual_src and manual_src.exists():
            link_name = f"说明书（{matched_manufacturer}-{matched_model}）"
            link_path = dir_path / link_name

            # 如果已有同名链接/目录，先删除
            if link_path.exists() or link_path.is_symlink():
                if os.name == 'nt':
                    subprocess.run(["cmd", "/c", "rmdir", "/S", "/Q", str(link_path)], check=False)
                else:
                    if link_path.is_symlink():
                        link_path.unlink()
                    elif link_path.is_dir():
                        shutil.rmtree(link_path)
                    else:
                        link_path.unlink()

            # 创建符号链接（Windows 使用 junction）
            try:
                if os.name == 'nt':
                    subprocess.run(
                        ["cmd", "/c", "mklink", "/J", str(link_path), str(manual_src)],
                        check=True, capture_output=True,
                    )
                else:
                    link_path.symlink_to(manual_src, target_is_directory=True)
                manual_linked = True
                manual_link_name = link_name
            except Exception as exc:
                pass  # 说明书链接失败不影响其他资源

    # 5. 从定值单内容提取电压等级（用于筛选整定原则）
    voltage_level = 0
    voltage_match = re.search(r'(\d+)\s*[kK][vV]', all_text)
    if voltage_match:
        voltage_level = int(voltage_match.group(1))

    # 设备类型映射（英文 -> 中文，中文 -> 英文）
    device_type_cn_map = {
        "transformer": "变压器",
        "line": "线路",
        "bus": "母线",
        "breaker": "母联分段",
        "capacitor": "电容器",
        "reactor": "电抗器",
        "grounding_transformer": "接地变",
        "station_transformer": "站用变",
    }
    # 反向映射（中文 -> 英文）
    device_type_en_map = {v: k for k, v in device_type_cn_map.items()}

    # matched_device_type 可能是中文（如"线路保护"）或英文（如"line"）
    if matched_device_type in device_type_cn_map:
        device_type_cn = device_type_cn_map[matched_device_type]
    elif matched_device_type and matched_device_type.rstrip("保护") in device_type_en_map:
        # 处理 "线路保护" -> "线路" 的情况
        device_type_cn = matched_device_type.rstrip("保护")
    elif matched_device_type in device_type_en_map:
        device_type_cn = matched_device_type
    else:
        # 从工作区名称或定值单内容推断设备类型
        device_type_cn = ""
        device_type_keywords = ["变压器", "线路", "母线", "母联", "分段", "电容器", "电抗器", "接地变", "站用变"]
        for keyword in device_type_keywords:
            if keyword in all_text:
                device_type_cn = keyword
                break

    # 6. 链接相关整定原则（创建临时目录，只包含相关文件）
    resources_root = Path.home() / ".nanobot" / "agentplayground" / "resources"
    principles_src = resources_root / "principles"
    principles_link = dir_path / "整定原则"
    logger.info(f"link_manuals: principles_src={principles_src}, exists={principles_src.exists()}, device_type_cn={device_type_cn}, voltage_level={voltage_level}")

    if principles_src.exists():
        # 清理旧链接
        if principles_link.exists() or principles_link.is_symlink():
            if os.name == 'nt':
                subprocess.run(["cmd", "/c", "rmdir", "/S", "/Q", str(principles_link)], check=False)
            else:
                if principles_link.is_symlink():
                    principles_link.unlink()
                else:
                    shutil.rmtree(principles_link)

        # 筛选相关文件
        related_principles = []
        if device_type_cn:
            for f in principles_src.iterdir():
                if not f.is_file():
                    continue
                name = f.name
                # 匹配设备类型
                if device_type_cn not in name:
                    continue
                # 匹配电压等级（如果有）
                if voltage_level > 0:
                    if f"{voltage_level}kV" in name or f"{voltage_level}kv" in name.lower():
                        related_principles.append(f)
                    elif "通用" in name or "接地" in name:  # 通用规则和接地变规则总是包含
                        related_principles.append(f)
                else:
                    related_principles.append(f)

        # 如果没有匹配到相关文件，包含通用规则
        if not related_principles:
            for f in principles_src.iterdir():
                if f.is_file() and "通用" in f.name:
                    related_principles.append(f)

        # 创建临时目录存放相关文件的符号链接
        if related_principles:
            temp_principles_dir = Path.home() / ".nanobot" / "agentplayground" / "temp" / f"principles_{ws}"
            temp_principles_dir.mkdir(parents=True, exist_ok=True)
            for f in related_principles:
                link_file = temp_principles_dir / f.name
                if not link_file.exists():
                    try:
                        if os.name == 'nt':
                            subprocess.run(["cmd", "/c", "mklink", "/H", str(link_file), str(f)], check=True, capture_output=True)
                        else:
                            link_file.symlink_to(f)
                    except Exception:
                        pass

            try:
                if os.name == 'nt':
                    subprocess.run(
                        ["cmd", "/c", "mklink", "/J", str(principles_link), str(temp_principles_dir)],
                        check=True, capture_output=True,
                    )
                else:
                    principles_link.symlink_to(temp_principles_dir, target_is_directory=True)
            except Exception:
                pass

    # 6b. 链接校核报告模板（通用 + 设备专属）
    templates_src = resources_root / "templates"
    templates_link = dir_path / "校核报告模板"
    logger.info(f"link_manuals: templates_src={templates_src}, exists={templates_src.exists()}")

    if templates_src.exists():
        # 清理旧链接
        if templates_link.exists() or templates_link.is_symlink():
            if os.name == 'nt':
                subprocess.run(["cmd", "/c", "rmdir", "/S", "/Q", str(templates_link)], check=False)
            else:
                if templates_link.is_symlink():
                    templates_link.unlink()
                else:
                    shutil.rmtree(templates_link)

        # 筛选相关模板文件：通用模板 + 设备专属模板
        related_templates = []
        for f in templates_src.iterdir():
            if not f.is_file():
                continue
            name = f.name
            # 通用模板总是包含
            if "通用" in name:
                related_templates.append(f)
                continue
            # 设备专属模板：匹配设备类型（校核报告模板-{电压等级}{设备类型}.md）
            if device_type_cn and device_type_cn in name and name.startswith("校核报告模板-"):
                # 匹配电压等级（如果有）
                if voltage_level > 0:
                    if f"{voltage_level}kV" in name or f"{voltage_level}kv" in name.lower():
                        related_templates.append(f)
                    elif "接地" in name or "母线" in name or "母联" in name or "电容" in name or "电抗" in name or "站用" in name:
                        # 这些设备类型没有电压等级区分，直接包含
                        related_templates.append(f)
                else:
                    related_templates.append(f)

        # 创建临时目录存放相关文件的符号链接
        if related_templates:
            temp_templates_dir = Path.home() / ".nanobot" / "agentplayground" / "temp" / f"templates_{ws}"
            temp_templates_dir.mkdir(parents=True, exist_ok=True)
            for f in related_templates:
                link_file = temp_templates_dir / f.name
                if not link_file.exists():
                    try:
                        if os.name == 'nt':
                            subprocess.run(["cmd", "/c", "mklink", "/H", str(link_file), str(f)], check=True, capture_output=True)
                        else:
                            link_file.symlink_to(f)
                    except Exception:
                        pass

            try:
                if os.name == 'nt':
                    subprocess.run(
                        ["cmd", "/c", "mklink", "/J", str(templates_link), str(temp_templates_dir)],
                        check=True, capture_output=True,
                    )
                else:
                    templates_link.symlink_to(temp_templates_dir, target_is_directory=True)
            except Exception:
                pass

    # 7. 链接相关台账
    account_src = resources_root / "account"
    account_link = dir_path / "台账"

    if account_src.exists():
        # 清理旧链接
        if account_link.exists() or account_link.is_symlink():
            if os.name == 'nt':
                subprocess.run(["cmd", "/c", "rmdir", "/S", "/Q", str(account_link)], check=False)
            else:
                if account_link.is_symlink():
                    account_link.unlink()
                else:
                    shutil.rmtree(account_link)

        # 筛选相关台账文件
        related_accounts = []
        if device_type_cn:
            for f in account_src.iterdir():
                if not f.is_file():
                    continue
                name = f.name
                # 匹配设备类型
                if device_type_cn in name:
                    related_accounts.append(f)
            # 线路台账总是包含线路厂站对应表
            if device_type_cn == "线路":
                for f in account_src.iterdir():
                    if "厂站对应" in f.name:
                        related_accounts.append(f)

        # 创建临时目录存放相关文件的符号链接
        if related_accounts:
            temp_account_dir = Path.home() / ".nanobot" / "agentplayground" / "temp" / f"account_{ws}"
            temp_account_dir.mkdir(parents=True, exist_ok=True)
            for f in related_accounts:
                link_file = temp_account_dir / f.name
                if not link_file.exists():
                    try:
                        if os.name == 'nt':
                            subprocess.run(["cmd", "/c", "mklink", "/H", str(link_file), str(f)], check=True, capture_output=True)
                        else:
                            link_file.symlink_to(f)
                    except Exception:
                        pass

            try:
                if os.name == 'nt':
                    subprocess.run(
                        ["cmd", "/c", "mklink", "/J", str(account_link), str(temp_account_dir)],
                        check=True, capture_output=True,
                    )
                else:
                    account_link.symlink_to(temp_account_dir, target_is_directory=True)
            except Exception:
                pass

    return {
        "linked": manual_linked,
        "manufacturer": matched_manufacturer,
        "device_type": matched_device_type,
        "device_type_cn": device_type_cn,
        "voltage_level": voltage_level,
        "model": matched_model,
        "path": manual_link_name if manual_linked else "",
    }


@router.get("/workspaces/{ws}/events")
async def workspace_events(ws: str):
    """SSE endpoint for workspace file change events"""
    dir_path = _ws_path(ws)
    if not dir_path.exists():
        return Response(content=json.dumps({"error": "workspace not found"}), status_code=404, media_type="application/json")

    async def event_generator():
        last_mtime = _get_dir_mtime(dir_path)
        while True:
            await asyncio.sleep(2)
            current_mtime = _get_dir_mtime(dir_path)
            if current_mtime > last_mtime:
                last_mtime = current_mtime
                yield f"data: {json.dumps({'type': 'change', 'path': '', 'mtime': current_mtime})}\n\n"

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )
