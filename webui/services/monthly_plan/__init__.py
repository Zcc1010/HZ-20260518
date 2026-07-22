# -*- coding: utf-8 -*-
"""月度计划处理服务 — 搜索关键词、标记黄色行、生成汇总。"""
from __future__ import annotations

import os
import tempfile
from pathlib import Path
from typing import Any

import xlrd
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# 关键词列表
KEYWORDS = [
    "启动", "流变", "更换", "改造", "迁改", "通道", "扩容",
    "升高", "改接", "开断", "CT", "送电", "扩建",
]

YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")


def process_monthly_plan(file_path: str, output_dir: str | None = None) -> dict[str, Any]:
    """处理月度计划 Excel，搜索关键词并标记黄色行。

    Args:
        file_path: 输入 Excel 文件路径（.xls 或 .xlsx）
        output_dir: 输出目录，默认为输入文件所在目录

    Returns:
        {
            "total_rows": int,
            "matched_rows": int,
            "keywords_found": {keyword: count},
            "matched_data": [row_dict, ...],
            "marked_file_path": str,
            "headers": [str, ...],
        }
    """
    file_path = str(file_path)
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"文件不存在: {file_path}")

    ext = Path(file_path).suffix.lower()
    if ext == ".xls":
        return _process_xls(file_path, output_dir)
    elif ext == ".xlsx":
        return _process_xlsx(file_path, output_dir)
    else:
        raise ValueError(f"不支持的文件格式: {ext}，请上传 .xls 或 .xlsx 文件")


def _process_xls(file_path: str, output_dir: str | None) -> dict[str, Any]:
    """处理 .xls 格式文件。"""
    wb = xlrd.open_workbook(file_path, formatting_info=True)
    sheet = wb.sheet_by_index(0)

    if sheet.nrows < 3:
        raise ValueError("Excel 文件至少需要包含表头和一行数据")

    # 动态查找表头行（包含"工作内容"的行）
    header_row = -1
    work_content_col = -1
    for r in range(min(10, sheet.nrows)):  # 前10行内查找
        for c in range(sheet.ncols):
            val = str(sheet.cell_value(r, c)).strip()
            if "工作内容" in val:
                header_row = r
                work_content_col = c
                break
        if header_row >= 0:
            break

    if header_row < 0:
        raise ValueError('未找到"工作内容"列，请确认 Excel 文件格式')

    # 读取表头
    headers = [str(sheet.cell_value(header_row, c)).strip() for c in range(sheet.ncols)]

    # 扫描数据行（从表头行的下一行开始）
    matched_data: list[dict] = []
    keywords_found: dict[str, int] = {}
    matched_row_indices: set[int] = set()  # 0-based row index in sheet

    for row_idx in range(header_row + 1, sheet.nrows):
        cell_value = str(sheet.cell_value(row_idx, work_content_col))
        matched_keywords = []
        for kw in KEYWORDS:
            if kw in cell_value:
                matched_keywords.append(kw)
                keywords_found[kw] = keywords_found.get(kw, 0) + 1

        if matched_keywords:
            matched_row_indices.add(row_idx)
            row_data = {}
            for col_idx, header in enumerate(headers):
                val = sheet.cell_value(row_idx, col_idx)
                row_data[header] = str(val) if val else ""
            row_data["_匹配关键词"] = "、".join(matched_keywords)
            matched_data.append(row_data)

    # 生成标记后的 .xlsx 文件
    marked_path = _write_marked_xlsx(
        sheet, headers, matched_row_indices, file_path, output_dir, header_row
    )

    return {
        "total_rows": sheet.nrows - header_row - 1,  # excluding header
        "matched_rows": len(matched_data),
        "keywords_found": keywords_found,
        "matched_data": matched_data,
        "marked_file_path": marked_path,
        "headers": headers,
    }


def _process_xlsx(file_path: str, output_dir: str | None) -> dict[str, Any]:
    """处理 .xlsx 格式文件。"""
    from openpyxl import load_workbook

    wb = load_workbook(file_path)
    ws = wb.active
    if ws is None or ws.max_row is None or ws.max_row < 2:
        raise ValueError("Excel 文件至少需要包含表头和一行数据")

    # 读取表头
    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value).strip() if cell.value else "")

    # 查找"工作内容"列
    work_content_col = None
    for i, h in enumerate(headers):
        if "工作内容" in h:
            work_content_col = i
            break
    if work_content_col is None:
        raise ValueError('未找到"工作内容"列，请确认 Excel 文件格式')

    # 扫描数据行
    matched_data: list[dict] = []
    keywords_found: dict[str, int] = {}
    matched_row_indices: set[int] = set()

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        cell_value = str(row[work_content_col].value or "")
        matched_keywords = []
        for kw in KEYWORDS:
            if kw in cell_value:
                matched_keywords.append(kw)
                keywords_found[kw] = keywords_found.get(kw, 0) + 1

        if matched_keywords:
            matched_row_indices.add(row_idx)
            row_data = {}
            for col_idx, header in enumerate(headers):
                val = row[col_idx].value if col_idx < len(row) else ""
                row_data[header] = str(val) if val else ""
            row_data["_匹配关键词"] = "、".join(matched_keywords)
            matched_data.append(row_data)

    # 标记黄色行
    for row_idx in matched_row_indices:
        for cell in ws[row_idx]:
            cell.fill = YELLOW_FILL

    # 保存标记文件
    if output_dir is None:
        output_dir = str(Path(file_path).parent)
    stem = Path(file_path).stem
    marked_path = str(Path(output_dir) / f"{stem}_标记.xlsx")
    wb.save(marked_path)

    return {
        "total_rows": ws.max_row - 1,
        "matched_rows": len(matched_data),
        "keywords_found": keywords_found,
        "matched_data": matched_data,
        "marked_file_path": marked_path,
        "headers": headers,
    }


def _write_marked_xlsx(
    sheet: xlrd.sheet.Sheet,
    headers: list[str],
    matched_row_indices: set[int],
    original_path: str,
    output_dir: str | None,
    header_row: int = 0,
) -> str:
    """将 xlrd 读取的数据写入 .xlsx 并标记黄色行。

    保留原始结构：标题行、空行、表头行、数据行。
    """
    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("无法创建工作表")

    # 写所有行（保留原始结构）
    for row_idx in range(sheet.nrows):
        for col_idx in range(sheet.ncols):
            cell = ws.cell(row=row_idx + 1, column=col_idx + 1)
            cell.value = sheet.cell_value(row_idx, col_idx)
            # 匹配行标记黄色
            if row_idx in matched_row_indices:
                cell.fill = YELLOW_FILL

    if output_dir is None:
        output_dir = str(Path(original_path).parent)
    stem = Path(original_path).stem
    marked_path = str(Path(output_dir) / f"{stem}_标记.xlsx")
    wb.save(marked_path)
    return marked_path


def format_result_for_agent(result: dict[str, Any]) -> str:
    """将处理结果格式化为 Agent 可读的文本。"""
    lines = []
    lines.append(f"## 月度计划处理结果\n")
    lines.append(f"- 总数据行数：{result['total_rows']}")
    lines.append(f"- 匹配行数：{result['matched_rows']}")

    if result["keywords_found"]:
        lines.append(f"\n### 关键词命中统计\n")
        lines.append("| 关键词 | 命中次数 |")
        lines.append("|--------|----------|")
        for kw, count in sorted(result["keywords_found"].items(), key=lambda x: -x[1]):
            lines.append(f"| {kw} | {count} |")

    if result["matched_data"]:
        lines.append(f"\n### 匹配明细\n")
        headers = result["headers"]
        # 简化表头（去掉过长的列）
        display_headers = ["单位", "设备名称", "工作内容", "工期", "开工时间", "完工时间", "工作类型", "风险等级", "_匹配关键词"]
        available = [h for h in display_headers if h in headers or h == "_匹配关键词"]

        lines.append("| " + " | ".join(available) + " |")
        lines.append("| " + " | ".join(["---"] * len(available)) + " |")
        for row in result["matched_data"]:
            vals = [row.get(h, "") for h in available]
            # 截断过长的内容
            vals = [v[:50] + "..." if len(v) > 50 else v for v in vals]
            lines.append("| " + " | ".join(vals) + " |")

    lines.append(f"\n标记后的文件已生成：{result['marked_file_path']}")
    return "\n".join(lines)
